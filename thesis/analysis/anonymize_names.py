#!/usr/bin/env python3
"""
Anonymize student names across all thesis and data files.

Replaces full student names with pseudonyms in "FirstName S." format.
Keeps instructor (Корнієнко Сергій) and institutional (ЦДПО) identifiers.

Usage:
    python anonymize_names.py          # Process all files
    python anonymize_names.py --dry    # Preview changes without writing
"""

import csv
import os
import re
import sys

# ──────────────────────────────────────────────────────────────────────
# MASTER MAPPING: all observed name variants → pseudonym
# Sorted into groups by canonical identity
# ──────────────────────────────────────────────────────────────────────

# Multi-word replacements (safe for global text replacement)
MULTIWORD_MAP = {
    # Артеменко Марія
    "Артеменко Марія": "Марія А.",
    # Бадюліна Ксенія (+ typos)
    "Бадюліна Ксенія": "Ксенія Б.",
    "Бдюліна Ксенія": "Ксенія Б.",
    # Байгуш Сергій
    "Байгуш Сергій": "Сергій Б.",
    # Бойко Марія
    "Бойко Марія": "Марія Бо.",
    # Бондар Аліна
    "Бондар Аліна": "Аліна Б.",
    # Борита Кароліна (+ typo)
    "Борита Кароліна": "Кароліна Б.",
    "Борита Каррліна": "Кароліна Б.",
    # Булахова Наталя
    "Булахова Наталя": "Наталя Б.",
    # Бєдарєва Анастасія
    "Бєдарєва Анастасія": "Анастасія Бє.",
    # Вовк Софія (multiple orderings and cases)
    "Вовк Софія": "Софія В.",
    "Софія Вовк": "Софія В.",
    "вовк софія": "Софія В.",
    # Городова Вероніка
    "Городова Вероніка": "Вероніка Г.",
    # Горбатков Назар
    "Горбатков Назар": "Назар Г.",
    # Гринчак Олександра
    "Гринчак Олександра": "Олександра Г.",
    # Грудницька Аріна
    "Грудницька Аріна": "Аріна Гр.",
    # Д'яконова Інна / Дьяконова Аня
    "Д'яконова Інна": "Інна Д.",
    "Дьяконова Аня": "Аня Д.",
    # Данюк Андрій (multiple cases)
    "ДАНЮК  АНДРІЙ": "Андрій Д.",
    "ДАНЮК АНДРІЙ": "Андрій Д.",
    "данюк андрій": "Андрій Д.",
    "Данюк Андрій": "Андрій Д.",
    # Дорошенко Наталя
    "Дорошенко Наталя": "Наталя Д.",
    # Єршов Олександр
    "Єршов Олександр": "Олександр Є.",
    # Зоріна Олександра / Саша (multiple variants)
    "Зоріна Саша": "Олександра З.",
    "Саша Зоріна": "Олександра З.",
    "Саша Зорина": "Олександра З.",
    "Зорина Саша": "Олександра З.",
    "Зоріна Олександра": "Олександра З.",
    # Зубко Аріна
    "Зубко Аріна": "Аріна З.",
    # Іванова Анастасія
    "Іванова Анастасія": "Анастасія І.",
    # Йожикова Катерина
    "Йожикова Катерина": "Катерина Й.",
    # Кайданович Ріна
    "Кайданович Ріна": "Ріна К.",
    # Каменєва Анастасія / Настя (+ Russian spelling)
    "Каменєва Анастасія": "Анастасія К.",
    "Каменєва Настя": "Настя К.",
    "Каменева Настя": "Настя К.",
    "Каменева Анастасія": "Анастасія К.",
    # Кістень Олександра (+ Russian spelling)
    "кістень олександра": "Олександра Кі.",
    "кистень олександра": "Олександра Кі.",
    "Кістень Олександра": "Олександра Кі.",
    # Клюка Ольга
    "Клюка Ольга": "Ольга К.",
    # Коптев Даниїл (multiple scripts)
    "Даниил Коптев": "Даниїл К.",
    "коптев даниил": "Даниїл К.",
    "Коптев Даниїл": "Даниїл К.",
    # Лотошенко Дмитро
    "Лотошенко Дмитро": "Дмитро Л.",
    # Матвієнко Діана
    "Матвієнко Діана": "Діана М.",
    # Махіня Єлизавета (multiple spellings)
    "Махіня Єлизавета": "Єлизавета М.",
    "Махиня Єлизавета": "Єлизавета М.",
    "махіня єлизавета": "Єлизавета М.",
    "махиня элизавета": "Єлизавета М.",
    "Махиня Элизавета": "Єлизавета М.",
    # Махецький Гліб (multiple spellings)
    "Гліб Махецький": "Гліб М.",
    "Глеб Махецкий": "Гліб М.",
    "Махецький Гліб": "Гліб М.",
    # Меньшова Аліса
    "Меньшова Аліса": "Аліса М.",
    # Мироненко Ігор (multiple orderings)
    "Мироненко Ігор": "Ігор М.",
    "Ігор Мироненко": "Ігор М.",
    # Михальчук Яна
    "Михальчук Яна": "Яна М.",
    # Міняйленко Анастасія / Настя (+ Russian spelling)
    "Міняйленко Анастасия": "Анастасія Мн.",
    "Міняйленко Анастасія": "Анастасія Мн.",
    "Міняйленко Настя": "Настя Мн.",
    # Новикова Анастасія (multiple spellings + emoji)
    "Новикова Анастасия \u2764\ufe0f\U0001f525": "Анастасія Н.",
    "Новикова Анастасия ❤️🔥": "Анастасія Н.",
    "Новикова Анастасия": "Анастасія Н.",
    "Новікова Анастасія": "Анастасія Н.",
    "Новикова Анастасія": "Анастасія Н.",
    # Олексашина Віра (multiple orderings/cases)
    "Олексашина Віра": "Віра О.",
    "Віра Олексашина": "Віра О.",
    "олексашина віра": "Віра О.",
    # Онищенко Маргарита
    "Онищенко Маргарита": "Маргарита О.",
    # Постнова Віолетта (multiple orderings/cases/typos)
    "Постнова Віолетта": "Віолетта П.",
    "Віолетта Постнова": "Віолетта П.",
    "віолетта постнова": "Віолетта П.",
    "По стнова": "Віолетта П.",
    # Рзагова Оля
    "Рзагова Оля": "Оля Р.",
    # Свінціцка Вікторія
    "Свінціцка Вікторія": "Вікторія С.",
    # Соломченко Поліна
    "Соломченко Поліна": "Поліна С.",
    # Сорокін Матвій (Russian spelling)
    "Сорокин Матвей": "Матвій С.",
    "Сорокін Матвій": "Матвій С.",
    # Сімакова Марія (+ typos and case variants)
    "Сімакова Марія": "Марія С.",
    "Сімакова марія": "Марія С.",
    "Cімакоа Марія": "Марія С.",
    # Талаш Данило (both orderings)
    "Талаш Данило": "Данило Т.",
    "Данило Талаш": "Данило Т.",
    # Тубольцева Карина
    "Тубольцева Карина": "Карина Т.",
    # Федоровська Анастасія (+ typo)
    "Федоровська Анастасія": "Анастасія Ф.",
    "Анастася Федоровська": "Анастасія Ф.",
    "Анастасія Федоровська": "Анастасія Ф.",
    # Хайтматова Аріна (both orderings)
    "Хайтматова Аріна": "Аріна Х.",
    "Аріна Хайтматова": "Аріна Х.",
    # Харченко Ксенія (both orderings)
    "Харченко Ксенія": "Ксенія Х.",
    "Ксенія Харченко": "Ксенія Х.",
    # Чернокур Анастасія (+ genitive typo)
    "Чернокур Анастасія": "Анастасія Ч.",
    "Чернокур Анастасії": "Анастасії Ч.",
    # Шехватова Поліна (+ Latin script)
    "Шехватова Поліна": "Поліна Ш.",
    "Shekhvatova Polina": "Поліна Ш.",
    # Шокалюк Микита
    "Шокалюк Микита": "Микита Ш.",
    # Ющик Валерія
    "Ющик Валерія": "Валерія Ю.",
    # Дима дивайн рапира (special display name)
    "Дима дивайн рапира": "Дмитро В.",
    # 🌸Алiнка🌸 (emoji display name → Бондар Аліна)
    "🌸Алiнка🌸": "Аліна Б.",
    "\U0001f338Алiнка\U0001f338": "Аліна Б.",
    # Ксюша<3 (special display name)
    "Ксюша<3": "Ксенія К.",
    # Анатолій Гібелінда (fake name used by Каменева Настя)
    "Анатолій Гібелінда": "Настя К.",
}

# Standalone name replacements — only applied to name columns/sender fields
# These are first-name-only identifiers in enrollment/chat data
STANDALONE_NAME_MAP = {
    "Олексашина": "Віра О.",
    "Постнова": "Віолетта П.",
    "постнова": "Віолетта П.",
    "Хайтматова": "Аріна Х.",
    "Михальчук": "Яна М.",
    "данюк": "Андрій Д.",
    "Байгуш": "Сергій Б.",
}

# First-name-only map (apply only in name fields, with word boundary check)
FIRST_NAME_MAP = {
    "Анастасия": "Анастасія Н.",  # Новикова in chat context
    "Мария": "Марія (батьки)",     # parent/relative in chat
    "Віолетта": "Віолетта П.",
    "Віолетта": "Віолетта П.",
    "Валена": "Валена",            # unknown identity, keep as-is
}

# English name replacements (for chapter3.tex)
ENGLISH_MAP = {
    "Svintsitska V.": "Viktoriia S.",
    "Bondar A.": "Alina B.",
    "Horodova V.": "Veronika H.",
    "Kamenieva A.": "Anastasiia K.",
    "Kliuka O.": "Olha K.",
    "Novykova A.": "Anastasiia N.",
    "Simakova M.": "Mariia S.",
    "Talash D.": "Danylo T.",
    "Bulakhova N.": "Nataliia B.",
    "Sasha Zorina": "Oleksandra Z.",
    "Khaitmatova A.": "Arina Kh.",
    "Violetta": "Violetta P.",
    "Yana": "Yana M.",
    "Nastya": "Nastia M.",
}

# ──────────────────────────────────────────────────────────────────────
# CORE FUNCTIONS
# ──────────────────────────────────────────────────────────────────────

def _build_sorted_replacements(mapping):
    """Sort replacement pairs by key length descending (longest first)."""
    return sorted(mapping.items(), key=lambda x: len(x[0]), reverse=True)


_MULTIWORD_SORTED = _build_sorted_replacements(MULTIWORD_MAP)
_ENGLISH_SORTED = _build_sorted_replacements(ENGLISH_MAP)


def anonymize(text):
    """Apply all multi-word name replacements to text (longest-first)."""
    for pattern, replacement in _MULTIWORD_SORTED:
        text = text.replace(pattern, replacement)
    return text


def anonymize_english(text):
    """Apply English name replacements to text (for chapter3.tex)."""
    for pattern, replacement in _ENGLISH_SORTED:
        text = text.replace(pattern, replacement)
    return text


def anonymize_name_field(name):
    """Anonymize a value from a name/sender column.

    Handles multi-word names, standalone surnames, and first-name-only entries.
    """
    name = name.strip()
    if not name:
        return name

    # Skip instructor and institutional names
    if name in ("Корнієнко Сергій", "Сергій Корнієнко", "ЦДПО", "system",
                "Корнієнко Сергій Сергійович"):
        return name

    # Try multi-word map first
    for pattern, replacement in _MULTIWORD_SORTED:
        if pattern in name:
            return name.replace(pattern, replacement)

    # Try standalone surname map
    for pattern, replacement in sorted(STANDALONE_NAME_MAP.items(),
                                       key=lambda x: len(x[0]), reverse=True):
        if name == pattern or name.lower() == pattern.lower():
            return replacement

    # Try first-name-only map
    for pattern, replacement in FIRST_NAME_MAP.items():
        if name == pattern:
            return replacement

    # Known first-name-only identifiers from enrollment_timeline
    first_name_enrollment = {
        "Яна": "Яна М.",
        "Микита": "Микита Ш.",
        "Віолетта": "Віолетта П.",
        "Вера": "Віра О.",
        "Дима": "Дмитро Л.",
        "Настя": "Настя Мн.",
        "Диана": "Діана М.",
        "Оля": "Оля Р.",
        "Глеб": "Гліб М.",
        "Катя": "Катя Й.",
    }
    if name in first_name_enrollment:
        return first_name_enrollment[name]

    return name


def anonymize_chat_text(text):
    """Anonymize a chat record line — handles ⁨Name⁩ markers and text."""
    # First apply multi-word replacements globally
    text = anonymize(text)

    # Replace names inside ⁨...⁩ markers (Telegram/Viber format)
    def replace_marker(m):
        inner = m.group(1)
        replaced = anonymize_name_field(inner)
        return f"\u2068{replaced}\u2069"

    text = re.sub(r'\u2068([^\u2069]+)\u2069', replace_marker, text)

    # Also handle ⁨...⁩ format (different Unicode chars sometimes used)
    text = re.sub(r'⁨([^⁩]+)⁩', lambda m: f'⁨{anonymize_name_field(m.group(1))}⁩', text)

    return text


# ──────────────────────────────────────────────────────────────────────
# FILE PROCESSING FUNCTIONS
# ──────────────────────────────────────────────────────────────────────

def process_csv(filepath, name_columns, dry=False):
    """Anonymize specified columns in a CSV file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Detect delimiter
    first_line = content.split('\n')[0]
    delimiter = ',' if ',' in first_line else '\t'

    rows = list(csv.reader(content.splitlines(), delimiter=delimiter))
    if not rows:
        return 0

    header = rows[0]
    changes = 0

    # Find column indices
    col_indices = []
    for col_name in name_columns:
        if isinstance(col_name, int):
            col_indices.append(col_name)
        elif col_name in header:
            col_indices.append(header.index(col_name))

    for i in range(1, len(rows)):
        for ci in col_indices:
            if ci < len(rows[i]):
                old_val = rows[i][ci]
                new_val = anonymize_name_field(old_val)
                if old_val != new_val:
                    rows[i][ci] = new_val
                    changes += 1

    if changes > 0 and not dry:
        with open(filepath, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerows(rows)

    return changes


def process_csv_text_column(filepath, text_columns, dry=False):
    """Anonymize names embedded in text columns of a CSV (e.g. system messages)."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    delimiter = ',' if ',' in content.split('\n')[0] else '\t'
    rows = list(csv.reader(content.splitlines(), delimiter=delimiter))
    if not rows:
        return 0

    header = rows[0]
    changes = 0

    col_indices = []
    for col_name in text_columns:
        if col_name in header:
            col_indices.append(header.index(col_name))

    for i in range(1, len(rows)):
        for ci in col_indices:
            if ci < len(rows[i]):
                old_val = rows[i][ci]
                new_val = anonymize(old_val)
                if old_val != new_val:
                    rows[i][ci] = new_val
                    changes += 1

    if changes > 0 and not dry:
        with open(filepath, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerows(rows)

    return changes


def process_chat_file(filepath, dry=False):
    """Anonymize a chat log file (chat records.txt or video chat.txt)."""
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    new_lines = []
    changes = 0
    for line in lines:
        new_line = anonymize_chat_text(line)
        if new_line != line:
            changes += 1
        new_lines.append(new_line)

    if changes > 0 and not dry:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.writelines(new_lines)

    return changes


def process_research_csv(filepath, dry=False):
    """Anonymize name column (usually col 3) in research_data CSV files."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    delimiter = ','
    rows = list(csv.reader(content.splitlines(), delimiter=delimiter))
    if len(rows) < 2:
        return 0

    header = rows[0]
    # Find the name column — typically contains "прізвище" or "ім'я"
    name_col = None
    for idx, col in enumerate(header):
        if 'прізвище' in col.lower() or "ім'я" in col.lower() or 'имя' in col.lower():
            name_col = idx
            break

    if name_col is None:
        return 0

    changes = 0
    for i in range(1, len(rows)):
        if name_col < len(rows[i]):
            old_val = rows[i][name_col]
            new_val = anonymize_name_field(old_val)
            if old_val != new_val:
                rows[i][name_col] = new_val
                changes += 1

    if changes > 0 and not dry:
        with open(filepath, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerows(rows)

    return changes


# ──────────────────────────────────────────────────────────────────────
# XLSX PROCESSING (optional — requires openpyxl)
# ──────────────────────────────────────────────────────────────────────

def process_xlsx(filepath, dry=False):
    """Anonymize name columns in XLSX files."""
    try:
        import openpyxl
    except ImportError:
        print(f"  SKIP (openpyxl not available): {filepath}")
        return 0

    wb = openpyxl.load_workbook(filepath)
    changes = 0
    for ws in wb.worksheets:
        # Find name columns by header
        name_cols = []
        if ws.max_row < 1:
            continue
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header and ('прізвище' in str(header).lower() or
                          "ім'я" in str(header).lower() or
                          'имя' in str(header).lower()):
                name_cols.append(col)

        for row in range(2, ws.max_row + 1):
            for col in name_cols:
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str):
                    new_val = anonymize_name_field(cell.value)
                    if new_val != cell.value:
                        cell.value = new_val
                        changes += 1

    if changes > 0 and not dry:
        wb.save(filepath)

    return changes


# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────

def main():
    dry = '--dry' in sys.argv

    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    repo = os.path.dirname(base)

    if dry:
        print("=== DRY RUN — no files will be modified ===\n")

    total = 0

    # 1. thesis/analysis/ CSVs
    print("--- thesis/analysis/ CSVs ---")

    f = os.path.join(base, 'analysis', 'individual_trajectories.csv')
    n = process_csv(f, ['name'], dry=dry)
    print(f"  {os.path.basename(f)}: {n} changes")
    total += n

    f = os.path.join(base, 'analysis', 'attendance_log.csv')
    n = process_csv(f, ['sender'], dry=dry)
    print(f"  {os.path.basename(f)}: {n} changes")
    total += n

    f = os.path.join(base, 'analysis', 'enrollment_timeline.csv')
    n = process_csv(f, ['user'], dry=dry)
    print(f"  {os.path.basename(f)}: {n} changes")
    total += n

    f = os.path.join(base, 'analysis', 'parsed_chat.csv')
    n1 = process_csv(f, ['sender', 'joined_user'], dry=dry)
    n2 = process_csv_text_column(f, ['text'], dry=dry)
    print(f"  {os.path.basename(f)}: {n1 + n2} changes")
    total += n1 + n2

    # 2. fragments/research_data/ CSVs
    print("\n--- fragments/research_data/ CSVs ---")
    rd = os.path.join(repo, 'fragments', 'research_data')
    if os.path.isdir(rd):
        for fname in sorted(os.listdir(rd)):
            if fname.endswith('.csv'):
                fp = os.path.join(rd, fname)
                n = process_research_csv(fp, dry=dry)
                print(f"  {fname}: {n} changes")
                total += n

    # 3. chat records.txt
    print("\n--- chat records.txt ---")
    f = os.path.join(repo, 'chat records.txt')
    if os.path.isfile(f):
        n = process_chat_file(f, dry=dry)
        print(f"  chat records.txt: {n} changes")
        total += n

    # 4. Video record_s/*/chat.txt
    print("\n--- Video record_s/*/chat.txt ---")
    vr = os.path.join(repo, 'Video record_s')
    if os.path.isdir(vr):
        for dirpath, dirnames, filenames in os.walk(vr):
            for fname in filenames:
                if fname == 'chat.txt':
                    fp = os.path.join(dirpath, fname)
                    n = process_chat_file(fp, dry=dry)
                    short = os.path.relpath(fp, repo)
                    print(f"  {short}: {n} changes")
                    total += n

    # 5. XLSX files (optional)
    print("\n--- XLSX files ---")
    for subdir in ['xlsx', 'new_data']:
        xd = os.path.join(rd, subdir) if os.path.isdir(rd) else None
        if xd and os.path.isdir(xd):
            for fname in sorted(os.listdir(xd)):
                if fname.endswith('.xlsx'):
                    fp = os.path.join(xd, fname)
                    n = process_xlsx(fp, dry=dry)
                    print(f"  {subdir}/{fname}: {n} changes")
                    total += n

    print(f"\n=== Total: {total} changes {'(dry run)' if dry else 'applied'} ===")


if __name__ == '__main__':
    main()
