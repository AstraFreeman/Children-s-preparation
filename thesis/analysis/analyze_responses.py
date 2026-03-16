#!/usr/bin/env python3
"""Analyze Google Forms response data from research_data CSVs and XLSX files."""

import csv
import re
import os
import statistics
from pathlib import Path
from collections import defaultdict

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "fragments" / "research_data"
XLSX_DIR = DATA_DIR / "xlsx"
NEW_DATA_DIR = DATA_DIR / "new_data"
OUTPUT_STATS = Path(__file__).resolve().parent / "response_stats.csv"
OUTPUT_INDIVIDUAL = Path(__file__).resolve().parent / "individual_trajectories.csv"
OUTPUT_MC = Path(__file__).resolve().parent / "modular_control_stats.csv"


def parse_score(score_str):
    """Parse score from format 'X / Y' and return (score, max_score)."""
    if not score_str:
        return None, None
    match = re.match(r"(\d+(?:[.,]\d+)?)\s*/\s*(\d+(?:[.,]\d+)?)", score_str.strip())
    if match:
        score = float(match.group(1).replace(",", "."))
        max_score = float(match.group(2).replace(",", "."))
        return score, max_score
    return None, None


def extract_seminar_number(filename):
    """Extract seminar number from filename like 'Семінар 25_0 - ...'"""
    match = re.search(r"Семінар\s*(\d+)", filename)
    if match:
        return int(match.group(1))
    return None


def find_name_col_index(headers):
    """Find the name column index from headers list."""
    for i, h in enumerate(headers):
        if not h:
            continue
        h_lower = h.lower().strip()
        if (h_lower.startswith("напишіть сво") or
            h_lower.startswith("напишіть своє") or
            h_lower.startswith("напишіть своєї") or  # typo variant
            h_lower.startswith("напишіть свої") or   # typo variant
            h_lower.startswith("напишіть свопрізвище") or  # typo variant (MC4)
            h_lower.startswith("ім'я") or
            h_lower.startswith("прізвище") or
            h_lower.startswith("пізвище") or  # typo: missing р
            h_lower == "name"):
            return i
    return None


def find_score_col_index(headers):
    """Find the score column index from headers list."""
    for i, h in enumerate(headers):
        if not h:
            continue
        h_stripped = h.strip()
        if h_stripped == "Результат" or h_stripped == "Score":
            return i
    return None


def find_timestamp_col_index(headers):
    """Find the timestamp column index from headers list."""
    for i, h in enumerate(headers):
        if not h:
            continue
        h_stripped = h.strip()
        if h_stripped == "Позначка часу" or h_stripped == "Timestamp":
            return i
    return None


def count_question_columns(headers, score_col, timestamp_col, name_col):
    """Count question columns to infer max score."""
    known_cols = {score_col, timestamp_col, name_col}
    # Also skip "Електронна адреса" and date columns
    skip_keywords = ["електронна адреса", "оберіть сьогоднішню", "напишіть сьогоднішню"]
    count = 0
    for i, h in enumerate(headers):
        if i in known_cols:
            continue
        if not h or not h.strip():
            continue
        h_lower = h.lower().strip()
        skip = False
        for kw in skip_keywords:
            if h_lower.startswith(kw):
                skip = True
                break
        if skip:
            continue
        count += 1
    return count


def read_csv_file(filepath):
    """Read a CSV and extract timestamps, names (if present), and scores."""
    responses = []
    encodings = ["utf-8", "utf-8-sig", "cp1251", "latin-1"]

    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc) as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames
                if not headers:
                    continue

                score_col = None
                timestamp_col = None
                name_col = None

                for h in headers:
                    h_stripped = h.strip()
                    h_lower = h_stripped.lower()
                    if h_stripped == "Результат" or h_stripped == "Score":
                        score_col = h
                    if h_stripped == "Позначка часу" or h_stripped == "Timestamp":
                        timestamp_col = h
                    if (h_lower.startswith("напишіть сво") or
                        h_lower.startswith("ім'я") or
                        h_lower.startswith("прізвище") or
                        h_lower.startswith("пізвище") or
                        h_lower == "name"):
                        name_col = h

                for row in reader:
                    score, max_score = None, None
                    if score_col and score_col in row:
                        score, max_score = parse_score(row[score_col])

                    timestamp = row.get(timestamp_col, "") if timestamp_col else ""
                    name = row.get(name_col, "") if name_col else ""

                    if score is not None:
                        responses.append({
                            "timestamp": timestamp,
                            "name": name.strip() if name else "",
                            "score": score,
                            "max_score": max_score,
                            "pct": round(score / max_score * 100, 1) if max_score else 0,
                        })

                return responses, headers
        except (UnicodeDecodeError, csv.Error):
            continue

    return [], []


def read_xlsx_file(filepath):
    """Read an XLSX file (active sheet only) and extract timestamps, names, and scores."""
    try:
        import openpyxl
    except ImportError:
        return []

    responses = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            wb.close()
            return responses

        headers = [str(h).strip() if h else "" for h in rows[0]]

        score_col = find_score_col_index(headers)
        timestamp_col = find_timestamp_col_index(headers)
        name_col = find_name_col_index(headers)

        # Infer max from question column count
        max_score_inferred = None
        q_count = count_question_columns(headers, score_col, timestamp_col, name_col)
        if q_count > 0:
            max_score_inferred = float(q_count)

        for row in rows[1:]:
            if score_col is not None and score_col < len(row):
                val = row[score_col]
                score = None
                max_score = max_score_inferred

                if isinstance(val, (int, float)):
                    score = float(val)
                elif isinstance(val, str) and "/" in val:
                    score, max_score = parse_score(val)

                if score is not None and max_score and max_score > 0:
                    timestamp = str(row[timestamp_col]) if timestamp_col is not None and timestamp_col < len(row) else ""
                    name = str(row[name_col]).strip() if name_col is not None and name_col < len(row) and row[name_col] else ""

                    responses.append({
                        "timestamp": timestamp,
                        "name": name,
                        "score": score,
                        "max_score": max_score,
                        "pct": round(score / max_score * 100, 1),
                    })

        wb.close()
    except Exception as e:
        print(f"  XLSX error: {e}")

    return responses


def read_xlsx_multisheet(filepath):
    """Read ALL sheets in an XLSX file, returning list of (sheet_name, responses)."""
    try:
        import openpyxl
    except ImportError:
        return []

    results = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            headers = [str(h).strip() if h else "" for h in rows[0]]

            score_col = find_score_col_index(headers)
            timestamp_col = find_timestamp_col_index(headers)
            name_col = find_name_col_index(headers)

            if score_col is None:
                continue

            # Infer max from question column count
            q_count = count_question_columns(headers, score_col, timestamp_col, name_col)
            max_score_inferred = float(q_count) if q_count > 0 else None

            sheet_responses = []
            for row in rows[1:]:
                if score_col < len(row):
                    val = row[score_col]
                    score = None
                    max_score = max_score_inferred

                    if isinstance(val, (int, float)) and val is not None:
                        score = float(val)
                    elif isinstance(val, str) and "/" in val:
                        score, max_score = parse_score(val)

                    if score is not None and max_score and max_score > 0:
                        timestamp = str(row[timestamp_col]) if timestamp_col is not None and timestamp_col < len(row) and row[timestamp_col] else ""
                        name = ""
                        if name_col is not None and name_col < len(row) and row[name_col]:
                            name = str(row[name_col]).strip()

                        sheet_responses.append({
                            "timestamp": timestamp,
                            "name": name,
                            "score": score,
                            "max_score": max_score,
                            "pct": round(score / max_score * 100, 1),
                        })

            if sheet_responses:
                results.append((sheet_name, sheet_responses))

        wb.close()
    except Exception as e:
        print(f"  XLSX multisheet error ({filepath.name}): {e}")

    return results


def compute_stats(sem_num, filename, responses):
    """Compute descriptive statistics from a list of responses."""
    scores = [r["score"] for r in responses]
    max_scores = [r["max_score"] for r in responses]
    pcts = [r["pct"] for r in responses]
    max_score = max_scores[0] if max_scores else 0

    return {
        "seminar": sem_num,
        "filename": filename,
        "N": len(scores),
        "max_possible": max_score,
        "mean_score": round(statistics.mean(scores), 2),
        "median_score": round(statistics.median(scores), 2),
        "sd_score": round(statistics.stdev(scores), 2) if len(scores) > 1 else 0,
        "min_score": min(scores),
        "max_score_achieved": max(scores),
        "mean_pct": round(statistics.mean(pcts), 1),
        "median_pct": round(statistics.median(pcts), 1),
    }


def aggregate_multisheet_seminar(sheet_results):
    """Aggregate multi-sheet seminar data.

    For each student, average their percentages across sub-blocks,
    then compute seminar stats from those per-student averages.
    For students without names, treat each response as an individual entry.
    """
    # Collect per-student scores across sheets
    named_scores = defaultdict(list)  # name -> list of pct
    unnamed_scores = []  # list of pct for unnamed respondents

    all_max_scores = []

    for sheet_name, responses in sheet_results:
        for r in responses:
            all_max_scores.append(r["max_score"])
            if r["name"]:
                named_scores[r["name"]].append(r["pct"])
            else:
                unnamed_scores.append(r["pct"])

    # Compute per-student averages for named students
    student_averages = []
    for name, pcts in named_scores.items():
        student_averages.append({
            "name": name,
            "pct": round(statistics.mean(pcts), 1),
        })

    # For unnamed, each entry is a separate "student"
    for pct in unnamed_scores:
        student_averages.append({
            "name": "",
            "pct": pct,
        })

    return student_averages, all_max_scores


def process_new_data(all_stats, all_individual, processed_seminars):
    """Process XLSX files from new_data/ directory (seminars 1-24)."""
    if not NEW_DATA_DIR.exists():
        print("  new_data/ directory not found, skipping")
        return

    xlsx_files = sorted(NEW_DATA_DIR.glob("Семінар*.xlsx"))
    if not xlsx_files:
        print("  No seminar XLSX files in new_data/")
        return

    print(f"\n  Processing {len(xlsx_files)} files from new_data/...")

    for filepath in xlsx_files:
        sem_num = extract_seminar_number(filepath.name)
        if sem_num is None:
            print(f"  Skipping (no seminar number): {filepath.name}")
            continue
        if sem_num in processed_seminars:
            print(f"  Skipping Seminar {sem_num} (already processed)")
            continue

        sheet_results = read_xlsx_multisheet(filepath)
        if not sheet_results:
            print(f"  No scoreable responses in: {filepath.name}")
            continue

        # Multi-sheet aggregation
        student_averages, all_max_scores = aggregate_multisheet_seminar(sheet_results)

        if not student_averages:
            continue

        pcts = [s["pct"] for s in student_averages]
        max_score_repr = all_max_scores[0] if all_max_scores else 0

        # Compute synthetic scores from pcts for stats
        stat = {
            "seminar": sem_num,
            "filename": filepath.name,
            "N": len(student_averages),
            "max_possible": max_score_repr,
            "mean_score": 0,
            "median_score": 0,
            "sd_score": round(statistics.stdev(pcts), 2) if len(pcts) > 1 else 0,
            "min_score": 0,
            "max_score_achieved": 0,
            "mean_pct": round(statistics.mean(pcts), 1),
            "median_pct": round(statistics.median(pcts), 1),
        }
        all_stats.append(stat)

        # Track individual trajectories
        for s in student_averages:
            if s["name"]:
                all_individual[s["name"]].append({
                    "seminar": sem_num,
                    "score": 0,
                    "max_score": 0,
                    "pct": s["pct"],
                    "type": "seminar",
                })

        n_sheets = len(sheet_results)
        print(f"  Seminar {sem_num}: N={stat['N']}, mean={stat['mean_pct']}%, "
              f"median={stat['median_pct']}% ({n_sheets} sheets)")
        processed_seminars.add(sem_num)


def process_modular_controls(all_individual):
    """Process modular control data from 'Модульний контроль ВСІ.xlsx'."""
    mc_file = NEW_DATA_DIR / "Модульний контроль ВСІ.xlsx"
    if not mc_file.exists():
        print("  Modular control file not found")
        return []

    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not available")
        return []

    mc_stats = []
    wb = openpyxl.load_workbook(mc_file, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        # Skip the summary sheet
        if "Зведена" in sheet_name:
            continue

        # Extract MC number
        mc_match = re.search(r"(\d+)", sheet_name)
        if not mc_match:
            continue
        mc_num = int(mc_match.group(1))

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h).strip() if h else "" for h in rows[0]]
        score_col = find_score_col_index(headers)
        timestamp_col = find_timestamp_col_index(headers)
        name_col = find_name_col_index(headers)

        if score_col is None:
            print(f"  MC {mc_num}: no score column found")
            continue

        q_count = count_question_columns(headers, score_col, timestamp_col, name_col)
        max_score_inferred = float(q_count) if q_count > 0 else None

        responses = []
        for row in rows[1:]:
            if score_col < len(row):
                val = row[score_col]
                score = None
                max_score = max_score_inferred

                if isinstance(val, (int, float)) and val is not None:
                    score = float(val)
                elif isinstance(val, str) and "/" in val:
                    score, max_score = parse_score(val)

                if score is not None and max_score and max_score > 0:
                    name = ""
                    if name_col is not None and name_col < len(row) and row[name_col]:
                        name = str(row[name_col]).strip()
                    timestamp = ""
                    if timestamp_col is not None and timestamp_col < len(row) and row[timestamp_col]:
                        timestamp = str(row[timestamp_col])

                    pct = round(score / max_score * 100, 1)
                    responses.append({
                        "name": name,
                        "score": score,
                        "max_score": max_score,
                        "pct": pct,
                        "timestamp": timestamp,
                    })

        if not responses:
            print(f"  MC {mc_num}: no valid responses")
            continue

        pcts = [r["pct"] for r in responses]
        max_score = responses[0]["max_score"]

        # Extract date from first timestamp
        date_str = ""
        for r in responses:
            if r["timestamp"]:
                date_match = re.match(r"(\d{4}-\d{2}-\d{2})", r["timestamp"])
                if date_match:
                    date_str = date_match.group(1)
                    break

        mc_stat = {
            "module": mc_num,
            "N": len(responses),
            "max_possible": max_score,
            "mean_pct": round(statistics.mean(pcts), 1),
            "median_pct": round(statistics.median(pcts), 1),
            "sd_pct": round(statistics.stdev(pcts), 2) if len(pcts) > 1 else 0,
            "date": date_str,
        }
        mc_stats.append(mc_stat)

        # Track individual trajectories
        for r in responses:
            if r["name"]:
                all_individual[r["name"]].append({
                    "seminar": f"MC{mc_num}",
                    "score": r["score"],
                    "max_score": r["max_score"],
                    "pct": r["pct"],
                    "type": "modular_control",
                })

        print(f"  MC {mc_num}: N={mc_stat['N']}, mean={mc_stat['mean_pct']}%, "
              f"median={mc_stat['median_pct']}%, date={date_str}")

    wb.close()
    return mc_stats


def main():
    if not DATA_DIR.exists():
        print(f"Error: Data directory not found: {DATA_DIR}")
        return

    csv_files = sorted(DATA_DIR.glob("*.csv"))

    all_stats = []
    all_individual = defaultdict(list)
    processed_seminars = set()

    # Phase 1: Process CSV files
    if csv_files:
        print("Processing CSV files...")
        for filepath in csv_files:
            sem_num = extract_seminar_number(filepath.name)
            if sem_num is None:
                print(f"  Skipping (no seminar number): {filepath.name}")
                continue

            responses, headers = read_csv_file(filepath)
            if not responses:
                print(f"  No scoreable responses in: {filepath.name}")
                continue

            stat = compute_stats(sem_num, filepath.name, responses)
            all_stats.append(stat)

            for r in responses:
                if r["name"]:
                    all_individual[r["name"]].append({
                        "seminar": sem_num,
                        "score": r["score"],
                        "max_score": r["max_score"],
                        "pct": r["pct"],
                        "type": "seminar",
                    })

            print(f"  Seminar {sem_num}: N={stat['N']}, mean={stat['mean_pct']}%, "
                  f"median={stat['median_pct']}%")
            processed_seminars.add(sem_num)
    else:
        print("No CSV files found in research_data/")

    # Phase 2: Process XLSX files from xlsx/ for seminars not yet covered
    if XLSX_DIR.exists():
        xlsx_files = sorted(XLSX_DIR.glob("*.xlsx"))
        print(f"\nProcessing {len(xlsx_files)} XLSX files from xlsx/...")
        for filepath in xlsx_files:
            sem_num = extract_seminar_number(filepath.name)
            if sem_num is None or sem_num in processed_seminars:
                continue

            responses = read_xlsx_file(filepath)
            if not responses:
                print(f"  No scoreable responses in XLSX: {filepath.name}")
                continue

            stat = compute_stats(sem_num, filepath.name, responses)
            all_stats.append(stat)

            for r in responses:
                if r["name"]:
                    all_individual[r["name"]].append({
                        "seminar": sem_num,
                        "score": r["score"],
                        "max_score": r["max_score"],
                        "pct": r["pct"],
                        "type": "seminar",
                    })

            print(f"  Seminar {sem_num} (XLSX): N={stat['N']}, mean={stat['mean_pct']}%, "
                  f"median={stat['median_pct']}%")
            processed_seminars.add(sem_num)

    # Phase 3: Process new_data/ XLSX files (seminars 1-24)
    print("\nProcessing new_data/ files...")
    process_new_data(all_stats, all_individual, processed_seminars)

    # Phase 4: Process modular controls
    print("\nProcessing modular controls...")
    mc_stats = process_modular_controls(all_individual)

    # Write seminar stats
    all_stats.sort(key=lambda x: x["seminar"])
    stat_fields = [
        "seminar", "filename", "N", "max_possible",
        "mean_score", "median_score", "sd_score",
        "min_score", "max_score_achieved",
        "mean_pct", "median_pct",
    ]
    with open(OUTPUT_STATS, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=stat_fields)
        writer.writeheader()
        writer.writerows(all_stats)

    # Write individual trajectories (with type column)
    indiv_rows = []
    for name, entries in sorted(all_individual.items()):
        for e in sorted(entries, key=lambda x: str(x["seminar"])):
            indiv_rows.append({
                "name": name,
                "seminar": e["seminar"],
                "score": e["score"],
                "max_score": e["max_score"],
                "pct": e["pct"],
                "type": e.get("type", "seminar"),
            })

    indiv_fields = ["name", "seminar", "score", "max_score", "pct", "type"]
    with open(OUTPUT_INDIVIDUAL, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=indiv_fields)
        writer.writeheader()
        writer.writerows(indiv_rows)

    # Write modular control stats
    if mc_stats:
        mc_stats.sort(key=lambda x: x["module"])
        mc_fields = ["module", "N", "max_possible", "mean_pct", "median_pct", "sd_pct", "date"]
        with open(OUTPUT_MC, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=mc_fields)
            writer.writeheader()
            writer.writerows(mc_stats)

    print(f"\nSummary: {len(all_stats)} seminars analyzed")
    print(f"Unique participants with names: {len(all_individual)}")
    print(f"Modular controls: {len(mc_stats)}")
    print(f"-> {OUTPUT_STATS}")
    print(f"-> {OUTPUT_INDIVIDUAL}")
    if mc_stats:
        print(f"-> {OUTPUT_MC}")


if __name__ == "__main__":
    main()
